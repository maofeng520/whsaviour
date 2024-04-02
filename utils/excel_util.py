import openpyxl
import json


class Excel_Util:
    @staticmethod
    def get_test_data_from_excel(file, sheet_name):
        # 1. 打开工作簿
        wb = openpyxl.load_workbook(filename=file)
        # 2. 获取sheet
        sh = wb[sheet_name]
        row = sh.max_row  # 行
        column = sh.max_column  # 列
        # 3. 读取数据
        data = []
        # 获取第一行拿到所有的key
        keys = []
        for i in range(1, column + 1):
            keys.append(sh.cell(1, i).value)  # cell单元格
        # 循环每一行，组成字典
        for i in range(2, row + 1):
            # 循环每一行的列
            # 搞一个临时变量用来存放每一行的数据
            temp = {}
            for j in range(1, column + 1):
                # 每个单元格而就是一个键值对
                # 获取对应列的键，注意列是1开头，索引是0开头
                # key = keys[j-1]
                # value = sh.cell(i, j).value
                # temp[key] = value
                temp[keys[j - 1]] = sh.cell(i, j).value
            try:
                temp['headers'] = json.loads(temp['headers'])
                temp['data'] = json.loads(temp['data'])
            except json.decoder.JSONDecodeError:
                raise ValueError('用例数据json格式错误！')
            # 把每一行数据形成的字典添加到data列表中
            data.append(temp)
            wb.save(file)
            wb.close()
        return data


read_excel = Excel_Util()

if __name__ == '__main__':
    pass
    # print(read_excel.get_test_data_from_excel("../testdata/excel/test_login.xlsx", "login"))
